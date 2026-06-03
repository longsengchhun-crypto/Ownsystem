import datetime
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.conf import settings
from django.db import models
from django.http import HttpResponse
from accounts.models import User
from services.models import Service
from bookings.models import Booking
from quotations.models import Quotation
from payments.models import Payment
from projects.models import Project, ProjectFile
from website.models import ContactMessage

# Excel and PDF Libraries
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

@login_required
def dashboard_home(request):
    if request.user.role == 'admin':
        # --- ADMIN DASHBOARD ---
        # Search & Filtering
        search_q = request.GET.get('q', '').strip()
        status_f = request.GET.get('status', '').strip()
        service_f = request.GET.get('service', '').strip()

        # Query sets
        bookings_qs = Booking.objects.select_related('client', 'service').order_by('-created_at')
        projects_qs = Project.objects.select_related('booking').order_by('-start_date')
        payments_qs = Payment.objects.select_related('booking').order_by('-submitted_at')
        messages_qs = ContactMessage.objects.all().order_by('-created_at')
        clients_qs = User.objects.filter(role='client').order_by('-created_at')

        # Apply Filters
        if search_q:
            bookings_qs = bookings_qs.filter(
                models.Q(booking_id__icontains=search_q) |
                models.Q(project_title__icontains=search_q) |
                models.Q(client__username__icontains=search_q) |
                models.Q(client__first_name__icontains=search_q) |
                models.Q(client__last_name__icontains=search_q)
            )
            projects_qs = projects_qs.filter(
                models.Q(booking__booking_id__icontains=search_q) |
                models.Q(booking__project_title__icontains=search_q)
            )
            payments_qs = payments_qs.filter(
                models.Q(payment_reference__icontains=search_q) |
                models.Q(booking__booking_id__icontains=search_q)
            )
            clients_qs = clients_qs.filter(
                models.Q(username__icontains=search_q) |
                models.Q(first_name__icontains=search_q) |
                models.Q(last_name__icontains=search_q)
            )

        if status_f:
            bookings_qs = bookings_qs.filter(status=status_f)
        if service_f:
            bookings_qs = bookings_qs.filter(service_id=service_f)

        # Dashboard Statistics
        total_clients = User.objects.filter(role='client').count()
        total_revenue = Payment.objects.filter(payment_status='Approved').aggregate(models.Sum('amount'))['amount__sum'] or 0.00
        pending_bookings = Booking.objects.filter(status='Pending').count()
        active_projects = Project.objects.filter(booking__status='In Progress').count()
        pending_payments = Booking.objects.filter(status='Awaiting Payment').count()

        services = Service.objects.filter(active=True)
        notification_status = {
            'email_ready': bool(settings.EMAIL_HOST_USER and settings.EMAIL_HOST_PASSWORD and settings.ADMIN_EMAIL_NOTIFICATION),
            'telegram_ready': bool(settings.TELEGRAM_BOT_TOKEN and settings.TELEGRAM_CHAT_ID),
            'admin_email': settings.ADMIN_EMAIL_NOTIFICATION,
            'telegram_chat_id': settings.TELEGRAM_CHAT_ID,
        }

        context = {
            'total_clients': total_clients,
            'total_revenue': total_revenue,
            'pending_bookings': pending_bookings,
            'active_projects': active_projects,
            'pending_payments': pending_payments,
            'bookings': bookings_qs[:10],
            'projects': projects_qs[:10],
            'payments': payments_qs[:10],
            'messages': messages_qs[:5],
            'clients': clients_qs[:10],
            'services': services,
            'search_q': search_q,
            'status_f': status_f,
            'service_f': service_f,
            'notification_status': notification_status,
        }
        return render(request, 'dashboard/admin_dashboard.html', context)
        
    else:
        # --- CLIENT DASHBOARD ---
        client_bookings = Booking.objects.filter(client=request.user).select_related('service').order_by('-created_at')
        client_projects = Project.objects.filter(booking__client=request.user).select_related('booking').order_by('-start_date')
        client_payments = Payment.objects.filter(booking__client=request.user).select_related('booking').order_by('-submitted_at')
        
        # Client Statistics
        total_projects = client_projects.count()
        active_projects = client_projects.filter(progress_percentage__lt=100).count()
        completed_projects = client_projects.filter(progress_percentage=100).count()
        pending_payments = client_bookings.filter(status='Awaiting Payment').count()

        # Files deliverables
        client_files = ProjectFile.objects.filter(project__booking__client=request.user).select_related('project__booking').order_by('-uploaded_at')

        context = {
            'bookings': client_bookings,
            'projects': client_projects,
            'payments': client_payments,
            'files': client_files,
            'total_projects': total_projects,
            'active_projects': active_projects,
            'completed_projects': completed_projects,
            'pending_payments': pending_payments,
        }
        return render(request, 'dashboard/client_dashboard.html', context)


@login_required
def export_excel_report(request):
    if request.user.role != 'admin':
        messages.error(request, "Access denied. Reports are for administrators only.")
        return redirect('dashboard:home')

    # Create Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Revenue & Booking Summary"

    # Styling Elements
    title_font = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Segoe UI", size=10, bold=True)
    regular_font = Font(name="Segoe UI", size=10)
    
    title_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Indigo Accent
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark Slate
    sub_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid") # Soft grey
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # 1. Sheet Title
    ws.merge_cells('A1:G1')
    ws['A1'] = "VANGUARD CREATIVE - BUSINESS PERFORMANCE REPORT"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # Date info
    ws['A3'] = f"Generated Date: {datetime.date.today().strftime('%B %d, %Y')}"
    ws['A3'].font = bold_font
    
    # 2. Key Metrics Summary
    ws['A5'] = "Key Business Metrics"
    ws['A5'].font = Font(name="Segoe UI", size=12, bold=True, color="1E293B")
    
    metrics = [
        ("Total Clients", User.objects.filter(role='client').count()),
        ("Total Revenue Generated", Payment.objects.filter(payment_status='Approved').aggregate(models.Sum('amount'))['amount__sum'] or 0.00),
        ("Completed Bookings", Booking.objects.filter(status='Completed').count()),
        ("Active In-Progress Projects", Project.objects.filter(booking__status='In Progress').count()),
    ]
    
    ws.append([]) # row 6
    ws.append(["Metric Name", "Value"]) # row 7
    for col in range(1, 3):
        cell = ws.cell(row=7, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left" if col==1 else "right")
    
    for m in metrics:
        ws.append([m[0], m[1]])
        row = ws.max_row
        ws.cell(row=row, column=1).font = regular_font
        ws.cell(row=row, column=2).font = bold_font
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
        if isinstance(m[1], float) or isinstance(m[1], int) and m[0] == "Total Revenue Generated":
            ws.cell(row=row, column=2).number_format = '$#,##0.00'
            
    # 3. Revenue Breakdown List
    start_row = ws.max_row + 3
    ws.cell(row=start_row, column=1, value="Revenue Transaction Logs").font = Font(name="Segoe UI", size=12, bold=True)
    
    headers = ["Payment Reference", "Booking ID", "Client Name", "Service Category", "Amount Verified", "Submission Date", "Status"]
    ws.append(headers)
    header_row = ws.max_row
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="right" if h == "Amount Verified" else "center" if h == "Status" or h == "Submission Date" else "left")

    payments = Payment.objects.select_related('booking__client', 'booking__service').all().order_by('-submitted_at')
    for p in payments:
        ws.append([
            p.payment_reference,
            p.booking.booking_id,
            p.booking.client.get_full_name() or p.booking.client.username,
            p.booking.service.name if p.booking.service else "Custom Service",
            float(p.amount),
            p.submitted_at.strftime('%Y-%m-%d %H:%M'),
            p.payment_status
        ])
        row_idx = ws.max_row
        for col_idx in range(1, 8):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = regular_font
            cell.border = thin_border
            if col_idx == 5:
                cell.number_format = '$#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in [6, 7]:
                cell.alignment = Alignment(horizontal="center")

    # Auto-adjust column width (skip MergedCell objects which have no column_letter)
    for col in ws.columns:
        max_len = 0
        col_letter = None
        for cell in col:
            from openpyxl.cell.cell import MergedCell
            if isinstance(cell, MergedCell):
                continue
            if col_letter is None:
                col_letter = cell.column_letter
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        if col_letter:
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="Vanguard_Revenue_Report.xlsx"'
    wb.save(response)
    return response


@login_required
def export_pdf_report(request):
    if request.user.role != 'admin':
        messages.error(request, "Access denied. Reports are for administrators only.")
        return redirect('dashboard:home')

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="Vanguard_Business_Report.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    story = []
    
    # Styles
    styles = getSampleStyleSheet()
    
    # Create unique styles to avoid collision crash
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=24,
        textColor=colors.HexColor("#4F46E5"), # Indigo
        spaceAfter=15
    )
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=12,
        textColor=colors.HexColor("#64748B"), # Slate Grey
        spaceAfter=25
    )
    h2_style = ParagraphStyle(
        'DocH2',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=14,
        textColor=colors.HexColor("#1E293B"), # Dark Slate
        spaceBefore=20,
        spaceAfter=10
    )
    body_style = ParagraphStyle(
        'DocBody',
        parent=styles['BodyText'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor("#334155")
    )
    th_style = ParagraphStyle(
        'TableHead',
        fontName='Helvetica-Bold',
        fontSize=10,
        textColor=colors.white,
        alignment=1 # Center
    )
    td_style = ParagraphStyle(
        'TableData',
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor("#334155")
    )

    story.append(Paragraph("Vanguard Creative", title_style))
    story.append(Paragraph(f"Monthly Business Report — Generated {datetime.date.today().strftime('%B %d, %Y')}", subtitle_style))
    
    # Summary Info Paragraph
    summary_text = (
        "This performance report documents user registrations, service reservation volumes, "
        "and verified revenue collections compiled for manual ABA transaction proofs. All data represents active records."
    )
    story.append(Paragraph(summary_text, body_style))
    story.append(Spacer(1, 15))

    # --- KPI Grid Table ---
    story.append(Paragraph("Key Metrics Summary", h2_style))
    
    total_clients = User.objects.filter(role='client').count()
    total_revenue = Payment.objects.filter(payment_status='Approved').aggregate(models.Sum('amount'))['amount__sum'] or 0.00
    pending_bookings = Booking.objects.filter(status='Pending').count()
    active_projects = Project.objects.filter(booking__status='In Progress').count()

    kpi_data = [
        [
            Paragraph("<b>Total Clients</b>", body_style),
            Paragraph(str(total_clients), body_style),
            Paragraph("<b>Total Revenue</b>", body_style),
            Paragraph(f"${total_revenue:,.2f}", body_style)
        ],
        [
            Paragraph("<b>Pending Bookings</b>", body_style),
            Paragraph(str(pending_bookings), body_style),
            Paragraph("<b>Active Projects</b>", body_style),
            Paragraph(str(active_projects), body_style)
        ]
    ]
    
    kpi_table = Table(kpi_data, colWidths=[130, 130, 130, 130])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#F8FAFC")),
        ('PADDING', (0,0), (-1,-1), 10),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor("#E2E8F0")),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 20))

    # --- Revenue Breakdown Table ---
    story.append(Paragraph("Verified Transaction Ledger", h2_style))
    
    payments = Payment.objects.select_related('booking__client').filter(payment_status='Approved').order_by('-submitted_at')[:15]
    
    table_data = [[
        Paragraph("<b>Reference</b>", th_style),
        Paragraph("<b>Booking ID</b>", th_style),
        Paragraph("<b>Client</b>", th_style),
        Paragraph("<b>Amount</b>", th_style),
        Paragraph("<b>Date Verified</b>", th_style)
    ]]
    
    for p in payments:
        table_data.append([
            Paragraph(p.payment_reference, td_style),
            Paragraph(p.booking.booking_id, td_style),
            Paragraph(p.booking.client.username, td_style),
            Paragraph(f"${p.amount:,.2f}", td_style),
            Paragraph(p.submitted_at.strftime('%Y-%m-%d'), td_style)
        ])
        
    payment_table = Table(table_data, colWidths=[120, 90, 110, 100, 110])
    payment_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1E293B")),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('PADDING', (0,0), (-1,-1), 6),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#F8FAFC")]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
    ]))
    
    story.append(payment_table)
    doc.build(story)
    return response
